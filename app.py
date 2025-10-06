import os
import math
import pandas as pd
import streamlit as st
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv
from google.cloud import vision
import re
import base64
import cv2
import numpy as np
from PIL import Image
import io
import tempfile

# ---- الإعدادات العامة / البيئة ----
load_dotenv()

USERNAME = "admin"
PASSWORD = "Moraqip@123"

st.set_page_config(page_title="المراقب الذكي", layout="wide")

# ---- إعداد Google Vision من secrets ----
def setup_google_vision():
    try:
        key_b64 = st.secrets["GOOGLE_VISION_KEY_B64"]
        key_bytes = base64.b64decode(key_b64)
        with open("google_vision.json", "wb") as f:
            f.write(key_bytes)
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_vision.json"
        return vision.ImageAnnotatorClient()
    except Exception as e:
        st.error(f"❌ لم يتم تحميل مفتاح Google Vision بشكل صحيح: {e}")
        return None

# ---- اتصال قاعدة البيانات ----
def get_conn():
    return psycopg2.connect(
        dbname=os.environ.get("DB_NAME"),
        user=os.environ.get("DB_USER"),
        password=os.environ.get("DB_PASSWORD"),
        host=os.environ.get("DB_HOST"),
        port=os.environ.get("DB_PORT"),
        sslmode=os.environ.get("DB_SSLMODE", "require")
    )

# ---- دالة تحويل الجنس ----
def map_gender(x):
    try:
        val = int(float(x))
        return "F" if val == 1 else "M"
    except:
        return "M"
# ---- تسجيل الدخول ----
# ---- تسجيل الدخول ----
def login():
    st.markdown(
        """
        <style>
        .login-container {
            display: flex;
            justify-content: center;
            align-items: flex-start; /* يرفع الصندوق لفوق */
            height: 100vh;
            padding-top: 10vh;       /* مسافة من فوق */
        }
        .login-box {
            background: #ffffff;
            padding: 1.5rem 2rem;
            border-radius: 12px;
            box-shadow: 0px 2px 12px rgba(0,0,0,0.1);
            text-align: center;
            width: 300px;
        }
        .stTextInput>div>div>input {
            text-align: center;
            font-size: 14px;
            height: 35px;
        }
        .stButton button {
            background: linear-gradient(90deg, #4e73df, #1cc88a);
            color: white;
            border-radius: 6px;
            padding: 0.4rem 0.8rem;
            font-size: 14px;
            font-weight: bold;
            transition: 0.2s;
            width: 100%;
        }
        .stButton button:hover {
            background: linear-gradient(90deg, #1cc88a, #4e73df);
            transform: scale(1.02);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown('<div class="login-container"><div class="login-box">', unsafe_allow_html=True)

    st.markdown("### 🔑 تسجيل الدخول")
    u = st.text_input("👤 اسم المستخدم", key="login_user")
    p = st.text_input("🔒 كلمة المرور", type="password", key="login_pass")

    # ✅ كبسة واحدة تكفي
    login_btn = st.button("🚀 دخول", key="login_btn")
    if login_btn:
        if u == USERNAME and p == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()   # إعادة تحميل الصفحة مباشرة
        else:
            st.error("❌ اسم المستخدم أو كلمة المرور غير صحيحة")

    st.markdown('</div></div>', unsafe_allow_html=True)


# ---- تحقق من حالة الجلسة ----
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

# ========================== الواجهة بعد تسجيل الدخول ==========================
st.title("📊 نينوى - البحث في سجلات الناخبين")
st.markdown("سيتم البحث في قواعد البيانات باستخدام الذكاء الاصطناعي 🤖")

# ====== تبويبات ======
tab_browse, tab_single, tab_file, tab_file_name_center, tab_count = st.tabs(
    ["📄 تصفّح السجلات", "🔍 بحث برقم", "📂 رفع ملف Excel", "📥 رفع Excel (اسم + رقم مركز)", "📦 عدّ البطاقات"]
)

# ----------------------------------------------------------------------------- #
# 1) 📄 تصفّح السجلات
# ----------------------------------------------------------------------------- #
with tab_browse:
    st.subheader("📄 تصفّح السجلات مع فلاتر")

    if "page" not in st.session_state:
        st.session_state.page = 1
    if "filters" not in st.session_state:
        st.session_state.filters = {"voter": "", "name": "", "center": ""}

    colf1, colf2, colf3, colf4 = st.columns([1,1,1,1])
    with colf1:
        voter_filter = st.text_input("🔢 رقم الناخب:", value=st.session_state.filters["voter"])
    with colf2:
        name_filter = st.text_input("🧑‍💼 الاسم:", value=st.session_state.filters["name"])
    with colf3:
        center_filter = st.text_input("🏫 مركز الاقتراع:", value=st.session_state.filters["center"])
    with colf4:
        page_size = st.selectbox("عدد الصفوف", [10, 20, 50, 100], index=1)

    if st.button("🔎 تطبيق الفلاتر"):
        st.session_state.filters = {
            "voter": voter_filter.strip(),
            "name": name_filter.strip(),
            "center": center_filter.strip(),
        }
        st.session_state.page = 1

    # --- بناء شروط البحث ---
    where_clauses, params = [], []
    if st.session_state.filters["voter"]:
        where_clauses.append('CAST("رقم الناخب" AS TEXT) ILIKE %s')
        params.append(f"%{st.session_state.filters['voter']}%")
    if st.session_state.filters["name"]:
        where_clauses.append('"الاسم الثلاثي" ILIKE %s')
        params.append(f"%{st.session_state.filters['name']}%")
    if st.session_state.filters["center"]:
        where_clauses.append('"اسم مركز الاقتراع" ILIKE %s')
        params.append(f"%{st.session_state.filters['center']}%")

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    count_sql = f'SELECT COUNT(*) FROM "naynawa" {where_sql};'
    offset = (st.session_state.page - 1) * page_size
    data_sql = f'''
        SELECT
            "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
            "اسم مركز الاقتراع","رقم مركز الاقتراع",
            "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
        FROM "naynawa"
        {where_sql}
        ORDER BY "رقم الناخب" ASC
        LIMIT %s OFFSET %s;
    '''

    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(count_sql, params)
            total_rows = cur.fetchone()[0]

        df = pd.read_sql_query(data_sql, conn, params=params + [page_size, offset])
        conn.close()

        if not df.empty:
            df = df.rename(columns={
                "رقم الناخب": "رقم الناخب",
                "الاسم الثلاثي": "الاسم",
                "الجنس": "الجنس",
                "هاتف": "رقم الهاتف",
                "رقم العائلة": "رقم العائلة",
                "اسم مركز الاقتراع": "مركز الاقتراع",
                "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                "المدينة": "المدينة",
                "رقم مركز التسجيل": "رقم مركز التسجيل",
                "اسم مركز التسجيل": "اسم مركز التسجيل",
                "تاريخ الميلاد": "تاريخ الميلاد"
            })
            df["الجنس"] = df["الجنس"].apply(map_gender)

        total_pages = max(1, math.ceil(total_rows / page_size))

        # ✅ عرض النتائج
        st.dataframe(df, use_container_width=True, height=500)

        c1, c2, c3 = st.columns([1,2,1])
        with c1:
            if st.button("⬅️ السابق", disabled=(st.session_state.page <= 1)):
                st.session_state.page -= 1
                st.experimental_rerun()
        with c2:
            st.markdown(f"<div style='text-align:center;font-weight:bold'>صفحة {st.session_state.page} من {total_pages}</div>", unsafe_allow_html=True)
        with c3:
            if st.button("التالي ➡️", disabled=(st.session_state.page >= total_pages)):
                st.session_state.page += 1
                st.experimental_rerun()

    except Exception as e:
        st.error(f"❌ خطأ أثناء التصفح: {e}")
# ----------------------------------------------------------------------------- #
# 2) 🔍 البحث برقم واحد
# ----------------------------------------------------------------------------- #
with tab_single:
    st.subheader("🔍 البحث برقم الناخب")
    voter_input = st.text_input("ادخل رقم الناخب:")
    if st.button("بحث"):
        try:
            conn = get_conn()
            query = """
                SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                       "اسم مركز الاقتراع","رقم مركز الاقتراع",
                       "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                FROM "naynawa" WHERE "رقم الناخب" LIKE %s
            """
            df = pd.read_sql_query(query, conn, params=(voter_input.strip(),))
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)

                st.dataframe(df, use_container_width=True, height=500)
            else:
                st.warning("⚠️ لم يتم العثور على نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")
# ----------------------------------------------------------------------------- #
# 3) 📂 رفع ملف Excel (معدل مع الأرقام غير الموجودة)
# ----------------------------------------------------------------------------- #
with tab_file:
    st.subheader("📂 البحث باستخدام ملف Excel")
    uploaded_file = st.file_uploader("📤 ارفع ملف (رقم الناخب)", type=["xlsx"])
    if uploaded_file and st.button("🚀 تشغيل البحث"):
        try:
            voters_df = pd.read_excel(uploaded_file, engine="openpyxl")
            voter_col = "رقم الناخب" if "رقم الناخب" in voters_df.columns else "VoterNo"
            voters_list = voters_df[voter_col].astype(str).tolist()

            conn = get_conn()
            placeholders = ",".join(["%s"] * len(voters_list))
            query = f"""
                SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                       "اسم مركز الاقتراع","رقم مركز الاقتراع",
                       "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                FROM "naynawa" WHERE "رقم الناخب" IN ({placeholders})
            """
            df = pd.read_sql_query(query, conn, params=voters_list)
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)

                df["رقم المندوب الرئيسي"] = ""
                df["الحالة"] = 0
                df["ملاحظة"] = ""
                df["رقم المحطة"] = 1

                df = df[["رقم الناخب","الاسم","الجنس","رقم الهاتف",
                         "رقم العائلة","مركز الاقتراع","رقم مركز الاقتراع","رقم المحطة",
                         "رقم المندوب الرئيسي","الحالة","ملاحظة"]]

                # ✅ إيجاد الأرقام غير الموجودة
                found_numbers = set(df["رقم الناخب"].astype(str).tolist())
                missing_numbers = [num for num in voters_list if num not in found_numbers]

                # عرض النتائج الموجودة
                st.dataframe(df, use_container_width=True, height=500)

                # ملف النتائج الموجودة
                output_file = "نتائج_البحث.xlsx"
                df.to_excel(output_file, index=False, engine="openpyxl")
                wb = load_workbook(output_file)
                wb.active.sheet_view.rightToLeft = True
                wb.save(output_file)
                with open(output_file, "rb") as f:
                    st.download_button("⬇️ تحميل النتائج", f,
                        file_name="نتائج_البحث.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # عرض وتحميل الأرقام غير الموجودة
                if missing_numbers:
                    st.warning("⚠️ الأرقام التالية لم يتم العثور عليها في قاعدة البيانات:")
                    st.write(missing_numbers)

                    missing_df = pd.DataFrame(missing_numbers, columns=["الأرقام غير الموجودة"])
                    miss_file = "missing_numbers.xlsx"
                    missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                    with open(miss_file, "rb") as f:
                        st.download_button("⬇️ تحميل الأرقام غير الموجودة", f,
                            file_name="الأرقام_غير_الموجودة.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.warning("⚠️ لا يوجد نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")

# ----------------------------------------------------------------------------- #
# 4) 📘 رفع Excel (الاسم + اسم مركز الاقتراع) — نسخة نهائية بخوارزمية تطابق ذكية
# ----------------------------------------------------------------------------- #
from rapidfuzz import fuzz
import re

with tab_file_name_center:
    st.subheader("📘 البحث الذكي باستخدام ملف Excel (الاسم + اسم مركز الاقتراع)")
    st.markdown("**الملف يجب أن يحتوي على عمودين:** `الاسم` و `اسم مركز الاقتراع`.")

    # ✅ دالة التطبيع العربي (للمقارنة الذكية)
    AR_DIACRITICS = str.maketrans('', '', ''.join([
        '\u0610','\u0611','\u0612','\u0613','\u0614','\u0615','\u0616','\u0617','\u0618','\u0619','\u061A',
        '\u064B','\u064C','\u064D','\u064E','\u064F','\u0650','\u0651','\u0652','\u0653','\u0654','\u0655',
        '\u0656','\u0657','\u0658','\u0659','\u065A','\u065B','\u065C','\u065D','\u065E','\u065F','\u0670'
    ]))
    def normalize_ar(text: str) -> str:
        if not text:
            return ""
        s = str(text)
        s = s.translate(AR_DIACRITICS)
        s = re.sub(r"[^\w\s]", "", s)
        s = s.replace("ـ", "").replace(" ", "").strip()
        s = (s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
             .replace("ؤ","و").replace("ئ","ي").replace("ى","ي").replace("ة","ه"))
        return s.lower()

    # ✅ دالة حساب التشابه
    def similarity(a, b):
        return fuzz.token_sort_ratio(normalize_ar(a), normalize_ar(b))

    file_nc = st.file_uploader("📤 ارفع ملف Excel يحتوي الاسم + اسم مركز الاقتراع", type=["xlsx"])
    run_nc = st.button("🚀 تشغيل البحث الذكي")

    if file_nc and run_nc:
        try:
            # ===== مرحلة 1: قراءة الملف =====
            progress = st.progress(0)
            status = st.empty()
            status.text("📂 جاري قراءة الملف...")

            xdf = pd.read_excel(file_nc, engine="openpyxl")
            progress.progress(10)

            # تنظيف أسماء الأعمدة
            def clean_col_name(c):
                return str(c).replace("\u200f", "").replace("\u200e", "").strip().lower()
            cleaned_cols = {clean_col_name(c): c for c in xdf.columns}

            name_col_candidates = ["الاسم", "الاسم الثلاثي", "name", "full name"]
            center_col_candidates = ["اسم مركز الاقتراع", "مركز الاقتراع", "polling center name", "center name"]

            def pick_col(cands):
                for c in cands:
                    if clean_col_name(c) in cleaned_cols:
                        return cleaned_cols[clean_col_name(c)]
                return None

            name_col = pick_col(name_col_candidates)
            center_col = pick_col(center_col_candidates)

            if not name_col or not center_col:
                st.error("❌ لم يتم العثور على الأعمدة المطلوبة.")
                st.stop()

            centers_list = xdf[center_col].dropna().astype(str).str.strip().tolist()
            unique_centers = sorted(list(set(centers_list)))

            if not unique_centers:
                st.warning("⚠️ لا توجد مراكز صالحة في الملف.")
                st.stop()

            status.text("🔌 الاتصال بقاعدة البيانات...")
            conn = get_conn()
            progress.progress(20)

            # ===== مرحلة 2: تحميل بيانات المراكز =====
            all_dfs = []
            batch_size = 100
            total_batches = (len(unique_centers) + batch_size - 1) // batch_size

            for i in range(total_batches):
                batch = unique_centers[i * batch_size : (i + 1) * batch_size]
                status.text(f"📦 تحميل بيانات المراكز {i+1}/{total_batches} ...")
                query = """
                    SELECT
                        "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                        "اسم مركز الاقتراع","رقم مركز الاقتراع",
                        "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                    FROM "naynawa"
                    WHERE "اسم مركز الاقتراع" = ANY(%s)
                """
                df_part = pd.read_sql_query(query, conn, params=(batch,))
                if not df_part.empty:
                    all_dfs.append(df_part)
                progress.progress(20 + int((i+1) / total_batches * 30))

            # إغلاق الاتصال بعد التحميل
            try:
                conn.close()
                if "db_conn" in st.session_state:
                    del st.session_state.db_conn
            except:
                pass

            if not all_dfs:
                st.warning("⚠️ لم يتم العثور على أي سجلات للمراكز.")
                st.stop()

            db_df = pd.concat(all_dfs, ignore_index=True)
            status.text(f"✅ تم تحميل {len(db_df)} سجل من قاعدة البيانات.")
            progress.progress(60)

            # ===== مرحلة 3: تطبيع الأسماء =====
            status.text("🧠 جاري تطبيع البيانات وتحليل الأسماء...")
            db_df["__norm_name"] = db_df["الاسم الثلاثي"].apply(normalize_ar)
            db_df["__norm_center"] = db_df["اسم مركز الاقتراع"].apply(normalize_ar)
            xdf["__norm_name"] = xdf[name_col].apply(normalize_ar)
            xdf["__norm_center"] = xdf[center_col].apply(normalize_ar)
            progress.progress(70)

            # ===== مرحلة 4: المطابقة الذكية =====
            status.text("🔎 جاري مطابقة الأسماء...")
            results = []
            total = len(xdf)
            for idx, row in xdf.iterrows():
                in_name = row["__norm_name"]
                in_center = row["__norm_center"]
                orig_name = row[name_col]
                orig_center = row[center_col]

                subset = db_df[db_df["__norm_center"] == in_center]
                if subset.empty:
                    results.append({
                        "الاسم (من الملف)": orig_name,
                        "اسم مركز الاقتراع (من الملف)": orig_center,
                        "الاسم في القاعدة": "—",
                        "درجة التطابق": 0,
                        "رقم الناخب": "",
                        "مركز الاقتراع": "",
                        "الحالة": 0
                    })
                else:
                    subset["درجة التطابق"] = subset["__norm_name"].apply(lambda x: similarity(in_name, x))
                    best = subset.sort_values("درجة التطابق", ascending=False).iloc[0]
                    results.append({
                        "الاسم (من الملف)": orig_name,
                        "اسم مركز الاقتراع (من الملف)": orig_center,
                        "الاسم في القاعدة": best["الاسم الثلاثي"],
                        "درجة التطابق": round(best["درجة التطابق"], 2),
                        "رقم الناخب": best["رقم الناخب"],
                        "مركز الاقتراع": best["اسم مركز الاقتراع"],
                        "الحالة": 0
                    })

                if idx % 50 == 0:
                    progress.progress(70 + int(idx / total * 25))

            progress.progress(95)
            status.text("🧾 تجهيز الملف النهائي...")

            # ===== مرحلة 5: تجهيز النتائج =====
            res_df = pd.DataFrame(results)
            res_df = res_df[[
                "رقم الناخب","الاسم (من الملف)","الاسم في القاعدة","درجة التطابق",
                "اسم مركز الاقتراع (من الملف)","مركز الاقتراع","الحالة"
            ]]

            # حفظ الإكسل
            out_file = "نتائج_التطابق_الذكي.xlsx"
            res_df.to_excel(out_file, index=False, engine="openpyxl")
            wb = load_workbook(out_file)
            wb.active.sheet_view.rightToLeft = True
            wb.save(out_file)
            with open(out_file, "rb") as f:
                st.download_button("⬇️ تحميل النتائج التفصيلية", f,
                    file_name="نتائج_التطابق_الذكي.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            progress.progress(100)
            status.text("🎯 تم إكمال التحليل بنجاح ✅")

            # عرض النتائج في الجدول
            st.dataframe(res_df, use_container_width=True, height=500)

        except Exception as e:
            st.error(f"❌ خطأ أثناء تنفيذ البحث: {e}")

# ----------------------------------------------------------------------------- #
# 5) 📦 عدّ البطاقات (أرقام 8 خانات) + بحث في القاعدة + قائمة الأرقام غير الموجودة
# ----------------------------------------------------------------------------- #
with tab_count:
    st.subheader("📦 عدّ البطاقات (أرقام 8 خانات) — بحث في القاعدة + الأرقام غير الموجودة")

    imgs_count = st.file_uploader(
        "📤 ارفع صور الصفحات (قد تحتوي أكثر من بطاقة)",
        type=["jpg","jpeg","png"],
        accept_multiple_files=True,
        key="ocr_count"
    )

    if imgs_count and st.button("🚀 عدّ البطاقات والبحث"):
        client = setup_google_vision()
        if client is None:
            st.error("❌ خطأ في إعداد Google Vision.")
        else:
            all_numbers, number_to_files, details = [], {}, []

            for img in imgs_count:
                try:
                    content = img.read()
                    image = vision.Image(content=content)
                    response = client.text_detection(image=image)
                    texts = response.text_annotations
                    full_text = texts[0].description if texts else ""

                    # استخراج أرقام مكونة من 8 خانات فقط
                    found_numbers = re.findall(r"\b\d{8}\b", full_text)
                    for n in found_numbers:
                        all_numbers.append(n)
                        number_to_files.setdefault(n, set()).add(img.name)

                    details.append({
                        "اسم الملف": img.name,
                        "عدد البطاقات (أرقام 8 خانات)": len(found_numbers),
                        "الأرقام المكتشفة (8 خانات فقط)": ", ".join(found_numbers) if found_numbers else "لا يوجد"
                    })

                except Exception as e:
                    st.warning(f"⚠️ خطأ أثناء معالجة صورة {img.name}: {e}")

            total_cards = len(all_numbers)
            unique_numbers = sorted(list(set(all_numbers)))

            st.success("✅ تم الاستخراج الأولي للأرقام")

            # ----------------- بحث في قاعدة البيانات عن الأرقام الموجودة -----------------
            found_df = pd.DataFrame()
            missing_list = []
            if unique_numbers:
                try:
                    conn = get_conn()
                    placeholders = ",".join(["%s"] * len(unique_numbers))
                    query = f"""
                        SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                               "اسم مركز الاقتراع","رقم مركز الاقتراع",
                               "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                        FROM "naynawa" WHERE "رقم الناخب" IN ({placeholders})
                    """
                    found_df = pd.read_sql_query(query, conn, params=unique_numbers)
                    conn.close()

                    if not found_df.empty:
                        found_df = found_df.rename(columns={
                            "رقم الناخب": "رقم الناخب",
                            "الاسم الثلاثي": "الاسم",
                            "الجنس": "الجنس",
                            "هاتف": "رقم الهاتف",
                            "رقم العائلة": "رقم العائلة",
                            "اسم مركز الاقتراع": "مركز الاقتراع",
                            "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                            "المدينة": "المدينة",
                            "رقم مركز التسجيل": "رقم مركز التسجيل",
                            "اسم مركز التسجيل": "اسم مركز التسجيل",
                            "تاريخ الميلاد": "تاريخ الميلاد"
                        })
                        found_df["الجنس"] = found_df["الجنس"].apply(map_gender)

                    found_numbers_in_db = set(found_df["رقم الناخب"].astype(str).tolist()) if not found_df.empty else set()
                    for n in unique_numbers:
                        if n not in found_numbers_in_db:
                            files = sorted(list(number_to_files.get(n, [])))
                            missing_list.append({"رقم_الناخب": n, "المصدر(الصور)": ", ".join(files)})
                except Exception as e:
                    st.error(f"❌ خطأ أثناء البحث في قاعدة البيانات: {e}")
            else:
                st.info("ℹ️ لم يتم العثور على أي أرقام مكوّنة من 8 خانات في الصور المرفوعة.")

            # ----------------- عرض النتائج للمستخدم -----------------
            st.markdown("### 📊 ملخص الاستخراج")
            st.metric("إجمالي الأرقام (مع التكرار)", total_cards)
            st.metric("إجمالي الأرقام الفريدة (8 خانات)", len(unique_numbers))
            st.metric("عدد الصور المرفوعة", len(imgs_count))

            st.markdown("### 🔎 بيانات الناخبين (الموجودة في قاعدة البيانات)")
            if not found_df.empty:
                st.dataframe(found_df, use_container_width=True, height=400)
                out_found = "found_voters.xlsx"
                found_df.to_excel(out_found, index=False, engine="openpyxl")
                with open(out_found, "rb") as f:
                    st.download_button("⬇️ تحميل بيانات الناخبين الموجودة", f,
                        file_name="بيانات_الناخبين_الموجودين.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("⚠️ لم يتم العثور على أي مطابقات في قاعدة البيانات.")

            st.markdown("### ❌ الأرقام غير الموجودة في القاعدة (مع اسم الصورة)")
            if missing_list:
                missing_df = pd.DataFrame(missing_list)
                st.dataframe(missing_df, use_container_width=True)
                miss_file = "missing_numbers_with_files.xlsx"
                missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                with open(miss_file, "rb") as f:
                    st.download_button("⬇️ تحميل الأرقام غير الموجودة مع المصدر", f,
                        file_name="الأرقام_غير_الموجودة_مع_المصدر.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.success("✅ لا توجد أرقام مفقودة (كل الأرقام الموجودة تم إيجادها في قاعدة البيانات).")
